"""incremental_update.py

Usage:
    incremental_update.py [-h] GP_SPREADSHEET_PATH MONTHLY_UPDATE_PATH OUTPUT_PATH FIRST_SEEN

This script is rather fragile - it assumes a particular format of the party spreadsheet and the incremental register
updates. It will break if either changes.
"""
from typing import Literal, Optional
from docopt import docopt
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import csv
from dataclasses import dataclass
import re

from towerhamlets.council import filter_out_post_code


@dataclass(frozen=True)
class Instruction:
    operation: Literal["create", "change", "delete"]
    prefix: str
    number: int
    suffix: int
    markers: Optional[str]
    date_of_birth: Optional[str]
    surname: str
    forename: str
    postcode: Optional[str]
    address_1: Optional[str]
    address_2: Optional[str]
    address_3: Optional[str]
    address_4: Optional[str]


def get_operation(row):
    if int(row["ElectorDeletedMonth"]) and int(row["ElectorCreatedMonth"]):
        return None
    if int(row["ElectorDeletedMonth"]):
        return "delete"
    if int(row["ElectorCreatedMonth"]):
        return "create"
    if int(row["ElectorChangedMonth"]):
        return "change"
    raise Exception("No operation")


def get_instructions(incremental_update_path: str):
    with open(incremental_update_path, encoding="windows-1252") as f:
        reader = csv.DictReader(f)

        for row in reader:
            operation = get_operation(row)
            if not operation:
                continue

            match = re.fullmatch(r"(\w\w\d)-(\d+)(/(\d+))?", row["ElectorNumber"])
            prefix = match.group(1)
            number = match.group(2)
            suffix = match.group(4)

            yield Instruction(
                operation=operation,
                prefix=prefix,
                number=int(number),
                suffix=int(suffix or 0),
                markers=row["MarkersRegisterText"],
                date_of_birth=row["ElectorDOB"],
                surname=row["ElectorSurname"],
                forename=row["ElectorForename"],
                postcode=row["PropertyPostCode"],
                address_1=filter_out_post_code(row["PropertyAddress1"]),
                address_2=filter_out_post_code(row["PropertyAddress2"]),
                address_3=filter_out_post_code(row["PropertyAddress3"]),
                address_4=filter_out_post_code(row["PropertyAddress4"]),
            )


def get_electors(worksheet: Worksheet) -> dict[tuple[str, int, int], list]:
    return {
        (prefix, number, suffix): data
        for prefix, number, suffix, *data in worksheet.iter_rows(
            min_row=2, values_only=True
        )
    }


def main(
    gp_spreadsheet_path: str,
    incremental_update_path: str,
    output_path: str,
    first_seen: str,
):
    workbook = openpyxl.load_workbook(gp_spreadsheet_path, read_only=True)
    worksheet = workbook.active
    headers = next(worksheet.iter_rows(values_only=True))
    electors = get_electors(worksheet)

    for instruction in get_instructions(incremental_update_path):
        if instruction.operation == "create":
            print(
                f"Creating {instruction.prefix}-{instruction.number}/{instruction.suffix}"
            )
            electors[(instruction.prefix, instruction.number, instruction.suffix)] = [
                instruction.markers,
                instruction.date_of_birth,
                instruction.surname,
                instruction.forename,
                instruction.address_1,
                instruction.address_2,
                instruction.address_3,
                instruction.address_4,
                instruction.postcode,
                first_seen,
            ]

        if instruction.operation == "change":
            print(
                f"Changing {instruction.prefix}-{instruction.number}/{instruction.suffix}"
            )
            elector = electors[
                (instruction.prefix, instruction.number, instruction.suffix)
            ]
            electors[(instruction.prefix, instruction.number, instruction.suffix)] = [
                instruction.markers,
                instruction.date_of_birth,
                instruction.surname,
                instruction.forename,
                instruction.address_1,
                instruction.address_2,
                instruction.address_3,
                instruction.address_4,
                instruction.postcode,
            ] + elector[9:]

        if instruction.operation == "delete":
            print(
                f"Deleting {instruction.prefix}-{instruction.number}/{instruction.suffix}"
            )
            del electors[(instruction.prefix, instruction.number, instruction.suffix)]

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(headers)

    for (prefix, number, suffix), data in sorted(electors.items()):
        worksheet.append([prefix, number, suffix, *data])

    workbook.save(output_path)


if __name__ == "__main__":
    arguments = docopt(__doc__)
    main(
        arguments["GP_SPREADSHEET_PATH"],
        arguments["MONTHLY_UPDATE_PATH"],
        arguments["OUTPUT_PATH"],
        arguments["FIRST_SEEN"],
    )
