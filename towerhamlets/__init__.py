import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


def get_electors(worksheet: Worksheet) -> dict[tuple[str, int, int], list]:
    return {
        (prefix, number, suffix): data
        for prefix, number, suffix, *data in worksheet.iter_rows(
            min_row=2, values_only=True
        )
        if prefix is not None
    }


def load_electors(path):
    workbook = openpyxl.load_workbook(path, read_only=True)
    worksheet = workbook.active
    headers = next(worksheet.iter_rows(values_only=True))
    electors = get_electors(worksheet)
    return (headers, electors)


def create_output(headers, electors, output_path):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(headers)

    for (prefix, number, suffix), data in sorted(electors.items()):
        worksheet.append([prefix, number, suffix, *data])

    workbook.save(output_path)
